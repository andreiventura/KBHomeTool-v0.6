from cryptography.fernet import Fernet
import base64

code = b"""

import pandas as pd
import os
import datetime
import lxml as lx
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Alignment, PatternFill

# ~~Current bug/s
    # 1. Correctly finds duplicates and copies duplicates only, but they get duplicated twice in the output file
    #    need to check what's causing it~~
# ~~If I can't fix the duplication bug, I'm reiterating to make a function that copies the following from both Sustain and Community sheets:
#   a. Placement name
#   b. Version
#   c. Start date
#   d. End date~~
# into a new sheet in the same excel file, find_and_copy_duplicates will now check duplicates in that new sheet instead (version 1 fix)
#   ---(version 2 fix) possible fix is to prune duplicates, within the report list, per version to prevent removing data with different
#      versions with the same placement name
#
#   things to figure out:
#       1. ~~copy dates as dates instead of values~~
#       2. ~~copy columns in specific order, i.e. Placement > version > start date > end date~~
#       3. ~~check date range, any start date after current date = do not duplicate, any end date after current date = do not duplicate~~
#       4. ~~custom formatting~~
#       5. ~~add "no issues" in cells after every row~~
#       6. get sum of total ads according to version (for checking if anything has 0 total ads)
#       7. ~~final output sheet will result after finishing duplicates and checking for errors and callouts~~
#          final output sheet will not have total ads number at the end
#def copy_sheet_contents

# Checks and copies duplicates ==========================================================================================================
def first_pair():
    def find_and_copy_duplicates(file1, file2, output_file):
        # Read Excel files into pandas DataFrames, takes data from specific columns
        df1 = pd.read_excel(file1, usecols='B,D,F')
        df2 = pd.read_excel(file2, sheet_name=['Sustain', 'Community'], index_col=None, usecols='B,D,E,M', skiprows=7)

        cdf = pd.concat(df2.values())
        cdf = cdf[cdf["Version Name"] != '']
        cdf = cdf.dropna()

        # Find duplicates based on all columns
        duplicates = pd.merge(df1, cdf, how='inner', left_on='Version', right_on='Version Name')
        duplicates = duplicates[duplicates['Version'] != 'No Version']
        duplicates.drop(['Placement Name', 'Version Name'], axis=1, inplace=True)
        
        # Write the duplicates to a new Excel file
        duplicates.to_excel(output_file, index=False)
        print(f'Duplicates found and saved to {output_file}')

    # Removes duplicates within versions ====================================================================================================
    def pruner(output_file):
        df = pd.read_excel(output_file, sheet_name='Sheet1')
        result_df = df.drop_duplicates(subset=['Placement','Version'], keep='last')
        #result_df = df.drop_duplicates(subset=['Placement','Start Date'], keep='last')

        # Sort in ascending order
        result_df.sort_values(by=['Version', 'Placement'], ascending=[True, True], inplace=True, ignore_index=True)

        # Second version of pruner removes duplicates within the same sheet
        my_file_writer = pd.ExcelWriter(output_file, engine='xlsxwriter')
        result_df.to_excel(my_file_writer, sheet_name='Sheet1', index=False, index_label='Campaign', na_rep='NaN')

        my_file_writer._save()
        
    # Adjusts column widths, and format date in columns automatically =======================================================================
    def format_columns(output_file):
        #------ Column width adjuster ------#
        #Read output file with pandas
        dfadj = pd.read_excel(output_file)

        #Assign pandas to make changes to output file
        writer = pd.ExcelWriter(output_file, engine='xlsxwriter')

        def merge_values(row):
            if row['Issue']=="Running before start date":  # replace 'Condition' with your actual condition
                # return row['Issue1'] + ' ' + row['Issue2']  # Merge values from columns A and B
                return row['Issue']
            elif row['Issue2']=="Running after end date":
                return row['Issue2']
            elif row['Issue']!="Running before start date" and row['Issue2']!="Running after end date":
                return row['Issue2']  # Keep only the value from column A if condition is not met
            
        # Apply the function to create a new column 'Merged'
        dfadj['Issue'] = dfadj.apply(merge_values, axis=1)
        dfadj.drop(['Issue2'], axis=1, inplace=True)

        # YYYY-mm-dd date formatter to mm-dd-YYYY
        dfadj['End Date']=pd.to_datetime(dfadj['End Date']).dt.strftime('%m/%d/%Y')
        dfadj['Start Date']=pd.to_datetime(dfadj['Start Date']).dt.strftime('%m/%d/%Y')

        dfadj.to_excel(writer, sheet_name='Sheet1', index=False, na_rep='NaN')
        workbook  = writer.book
        worksheet = writer.sheets['Sheet1']
        
        for column in dfadj:
            column_length = max(dfadj[column].astype(str).map(len).max(), len(column))
            col_idx = dfadj.columns.get_loc(column)
            writer.sheets['Sheet1'].set_column(col_idx, col_idx, column_length)
        
        #------ Column width adjuster ------#
        merge_format = workbook.add_format({'align': 'center', 'valign': 'vcenter', 'border': 2})
        startCells = [1]
        for row in range(2,len(dfadj)+1):
            if (dfadj.loc[row-1,'Version'] != dfadj.loc[row-2,'Version']):
                startCells.append(row)
        lastRow = len(dfadj)

        for row in startCells:
            try:
                endRow = startCells[startCells.index(row)+1]-1
                if row == endRow:
                    worksheet.write(row, 2, dfadj.loc[row-1,'Version'], merge_format)
                else:
                    worksheet.merge_range(row, 2, endRow, 2, dfadj.loc[row-1,'Version'], merge_format)
            except IndexError:
                if row == lastRow:
                    worksheet.write(row, 2, dfadj.loc[row-1,'Version'], merge_format)
                else:
                    worksheet.merge_range(row, 2, lastRow, 2, dfadj.loc[row-1,'Version'], merge_format)
            
        writer._save()
    
    def date_toyear(output_file):
        dfadj = pd.read_excel(output_file)
        writer = pd.ExcelWriter(output_file, engine='xlsxwriter')

        dfadj.to_excel(writer, sheet_name='Sheet1', index=False, na_rep='NaN')
        workbook  = writer.book
        worksheet = writer.sheets['Sheet1']
        date_format = workbook.add_format({'num_format': 'yyyy/mm/dd'})

        # Set the column widths and add a date format.
        worksheet.set_column('D:E', 15, date_format)
        writer._save()

    def compare_dates(output_file):
        # converts raw date value to yyyy-dd-mm and compares to current date
        df = pd.read_excel(output_file, sheet_name='Sheet1')
        writer = pd.ExcelWriter(output_file, engine='xlsxwriter')

        df['End Date']=df['End Date'].dt.strftime('%Y-%m-%d')
        df['Start Date']=df['Start Date'].dt.strftime('%Y-%m-%d')

        def isdate(x):
            if x <= datetime.date.today().strftime('%Y-%m-%d'):
                x = 'No issues'
            elif x > datetime.date.today().strftime('%Y-%m-%d'):
                x = 'Running before start date'
            return(x)
        
        def isoverdate(y):
            if y < datetime.date.today().strftime('%Y-%m-%d'):
                y = 'Running after end date'
            elif y >= datetime.date.today().strftime('%Y-%m-%d'):
                y = 'No issues'
            return(y)

        df['Issue'] = df['Start Date'].apply(isdate)
        df['Issue2'] = df['End Date'].apply(isoverdate)

        df.to_excel(writer, sheet_name='Sheet1', index=False, na_rep='NaN')
        writer._save()

    def merge_campaign(output_file):
        def merge_center_rows_by_column(ws, merge_column_index):
            unique_values = set()
            for cell in ws.iter_rows(min_row=2, min_col=merge_column_index, max_col=merge_column_index, values_only=True):
                unique_values.add(cell[0])

            for value in unique_values:
                merge_start = None
                for row in range(2, ws.max_row + 2):  # Adjusted the end condition
                    if ws.cell(row=row, column=merge_column_index).value == value:
                        if merge_start is None:
                            merge_start = row
                    elif merge_start is not None:
                        ws.merge_cells(start_row=merge_start, start_column=merge_column_index, end_row=row - 1, end_column=merge_column_index)
                        ws.cell(row=merge_start, column=merge_column_index).alignment = openpyxl.styles.Alignment(horizontal='center')
                        merge_start = None

                if merge_start is not None:
                    ws.merge_cells(start_row=merge_start, start_column=merge_column_index, end_row=ws.max_row, end_column=merge_column_index)
                    ws.cell(row=merge_start, column=merge_column_index).alignment = openpyxl.styles.Alignment(horizontal='center')

        workbook = openpyxl.load_workbook(output_file)
        worksheet = workbook.active

        merge_center_rows_by_column(worksheet, merge_column_index=1)

        workbook.save(output_file)

    def merge_columns(output_file):
        def get_merged_cell_value(ws, row, column):
            for merged_cell_range in ws.merged_cells.ranges:
                if merged_cell_range.min_row <= row <= merged_cell_range.max_row and merged_cell_range.min_col <= column <= merged_cell_range.max_col:
                    return ws.cell(row=merged_cell_range.min_row, column=merged_cell_range.min_col).value
            return ws.cell(row=row, column=column).value

        def merge_center_rows_by_columns(ws, merge_column_indices, unique_values_column_index):
            unique_values = set()
            for row in range(2, ws.max_row + 1):
                unique_value = get_merged_cell_value(ws, row, unique_values_column_index)
                unique_values.add(unique_value)

            for value in unique_values:
                merge_start = None
                for row in range(2, ws.max_row + 2):  # Adjusted the end condition
                    cell_value = get_merged_cell_value(ws, row, unique_values_column_index)
                    if row <= ws.max_row and cell_value == value:
                        if merge_start is None:
                            merge_start = row
                    elif merge_start is not None:
                        for col in merge_column_indices:
                            ws.merge_cells(start_row=merge_start, start_column=col, end_row=row - 1, end_column=col)
                            ws.cell(row=merge_start, column=col).alignment = openpyxl.styles.Alignment(horizontal='center')
                        merge_start = None

        workbook = openpyxl.load_workbook(output_file)
        worksheet = workbook.active

        merge_center_rows_by_columns(worksheet, merge_column_indices=[4, 5, 6], unique_values_column_index=3)

        workbook.save(output_file)
    
    def cell_counter(output_file):
        def get_merged_cell_value(ws, row, column):
            for merged_cell_range in ws.merged_cells.ranges:
                if merged_cell_range.min_row <= row <= merged_cell_range.max_row and merged_cell_range.min_col <= column <= merged_cell_range.max_col:
                    return ws.cell(row=merged_cell_range.min_row, column=merged_cell_range.min_col).value
            return ws.cell(row=row, column=column).value

        def merge_center_rows_by_columns(ws, merge_column_indices, unique_values_column_index):
            unique_values = set()
            for row in range(2, ws.max_row + 1):
                unique_value = get_merged_cell_value(ws, row, unique_values_column_index)
                unique_values.add(unique_value)

            for value in unique_values:
                merge_start = None
                row_counter = 0  # Initialize row counter
                highlighted_column = None  # Initialize highlighted column
                for row in range(2, ws.max_row + 2):  
                    cell_value = get_merged_cell_value(ws, row, unique_values_column_index)
                    if row <= ws.max_row and cell_value == value:
                        if merge_start is None:
                            merge_start = row
                        row_counter += 1  # Increment row counter
                    elif merge_start is not None:
                        for col in merge_column_indices:
                            ws.merge_cells(start_row=merge_start, start_column=col, end_row=row - 1, end_column=col)
                            ws.cell(row=merge_start, column=col).alignment = Alignment(horizontal='center')
                            # Insert count value into the merged cell
                            ws.cell(row=merge_start, column=col).value = row_counter

                            if 3 < row_counter < 5:
                                ws.cell(row=merge_start, column=8).value = "Missing placement/s"
                            elif 5 < row_counter < 8:
                                ws.cell(row=merge_start, column=8).value = "Missing placement/s"
                            elif 8 < row_counter < 10:
                                ws.cell(row=merge_start, column=8).value = "Missing placement/s"
                        # Conditional formatting based on row_counter value
                        if 3 < row_counter < 5:
                            highlight_color = "FFE5E5"  # Pink
                        elif 5 < row_counter < 8:
                            highlight_color = "FFE5E5"  # Pink
                        elif 8 < row_counter < 10:
                            highlight_color = "FFE5E5"  # Pink
                        else:
                            highlight_color = None
                        
                        if highlight_color:
                            for r in range(merge_start, row):
                                for c in range(1, ws.max_column + 1):
                                    ws.cell(row=r, column=c).fill = PatternFill(start_color=highlight_color, end_color=highlight_color, fill_type="solid")
                                highlighted_column = ws.cell(row=1, column=7).coordinate  # Get the cell coordinate of the highlighted row's column

                        merge_start = None
                        row_counter = 0  # Reset row counter

                if highlighted_column:
                    ws[highlighted_column] = 'Number of Placements'
                    ws.cell(row=1, column=8).value = 'Placement Issues'  # Add label for column 8
                    # Auto adjust width of the column
                    col_letter = highlighted_column[:1]  # Extract column letter from the coordinate
                    ws.column_dimensions[col_letter].width = max(len('Number of Placements') + 2, ws.column_dimensions[col_letter].width)  # Set width to the maximum between 'Highlighted' length and current width

                    # Auto adjust width of the column for 'Label for Column 8'
                    label_col_width = max(len('Placement Issues') + 2, ws.column_dimensions['H'].width)
                    ws.column_dimensions['H'].width = label_col_width

        # Assuming 'output_file' is defined elsewhere in your code
        workbook = openpyxl.load_workbook(output_file)
        worksheet = workbook.active

        merge_center_rows_by_columns(worksheet, merge_column_indices=[7], unique_values_column_index=3)

        workbook.save(output_file)
    
    def cell_highlighter(output_file):
        workbook = openpyxl.load_workbook(output_file)
        worksheet = workbook.active

        highlight_color = 'FFD1D1'

        # Iterate through each row and check the condition
        for row in worksheet.iter_rows(min_row=2, min_col=6, max_row=worksheet.max_row, max_col=6):
            cell = row[0]
            if cell.value != 'No issues':
                # Highlight the entire row
                for cell in row:
                    cell.fill = PatternFill(start_color=highlight_color, end_color=highlight_color, fill_type="solid")

        workbook.save(output_file)

    while True:
        while True:
            current_directory = os.getcwd()

            # Prompt user for file name
            file_name1 = input("Enter the name of the Report file (including extension): ")
            file_name2 = input("Enter the name of the Traffic sheet (including extension): ")
            output_file = input("Enter the name of the Output file (including extension): ")

            # Construct the full file path
            file1_path = os.path.join(current_directory, file_name1)
            file2_path = os.path.join(current_directory, file_name2)
            output_file_path = os.path.join(current_directory, output_file)

            try:
                # Check if file paths are blank
                if file_name1.strip() == '' or file_name2.strip() == '' or output_file.strip() == '':
                    raise ValueError("File name cannot be blank.")

                # Check if files exist
                if not os.path.isfile(file1_path):
                    raise FileNotFoundError(f"Report file '{file_name1}' not found.")
                if not os.path.isfile(file2_path):
                    raise FileNotFoundError(f"Traffic sheet '{file_name2}' not found.")

                print("Saving results to ", output_file)

                find_and_copy_duplicates(file1_path, file2_path, output_file_path)
                pruner(output_file_path)
                date_toyear(output_file_path)
                compare_dates(output_file_path)
                format_columns(output_file_path)
                merge_campaign(output_file_path)
                merge_columns(output_file_path)
                cell_counter(output_file_path)
                cell_highlighter(output_file_path)

                print("Results saved to", output_file)
                repeat = input("Do you want to process more files? (yes/no): ").lower()
                if repeat != 'yes':
                    break  # Exit the inner loop if the user does not want to repeat
                    

            except (ValueError, FileNotFoundError) as e:
                print("Error:", e)
                cancel = input("Do you want to cancel? (yes/no): ").lower()
                if cancel == 'yes':
                    print("Operation canceled.")
                    break  # Exit the loop if the user wants to cancel
        
        # Prompt user if they want to stop or repeat the process
        stop_or_repeat = input("Do you want to exit and close? (yes/no): ").lower()
        if stop_or_repeat == 'yes':
            print("Exiting the program.")
            break  # Exit the outer loop if the user wants to stop
first_pair()

"""

key = Fernet.generate_key()
encryption_type = Fernet(key)
encrypted_message = encryption_type.encrypt(code)

decrypted_message = encryption_type.decrypt(encrypted_message)

exec(decrypted_message)